import pandas as pd
from openpyxl import Workbook

color = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Color info_detail.xlsx')
sku = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Cabinet_detailxlsx.xlsx')

# 读取第一个 Excel 文件，提取指定列的数据
cName = pd.read_excel(color, sheet_name='All', usecols=['Name','Code'])

# 读取第二个 Excel 文件，提取指定列的数据
pSku= pd.read_excel(sku, sheet_name='Test', usecols=['SKU','T','E','B','A1','A2','A3'])

# colors = ['Brushed Aluminum', 'River Rock', 'Sheer Beauty', 'Fashionista', 'The Chameleon', 'Weekend Getaway', 'Winter Fun', 'Casting at First Light', 'Sugar on Ice', 'Sand Gladstone Oak', 'Grey-Beige Gladstone Oak', 'Brown Tossini Elm', 'Tobacco Gladstone Oak', 'Tobacco Halifax Oak', 'Black Halifax Oak', 'Natural Halifax Oak', 'White Halifax Oak', 'Pearl White HG', 'Winter Frost SM', 'Sun Grey HG', 'Sun Grey SM', 'Stone Grey HG', 'Stone Grey SM', 'Eclipse  HG', 'Eclipse  SM', 'Royal Blue HG', 'Royal Blue SM', 'Majestic HG', 'Majestic SM', 'Ida 01', 'Ida 03', 'Roble Muratti 01', 'Roble Muratti 04', 'Factory 01', 'Factory 02', 'Como Ash 01', 'Como Ash 03', 'Gris Nube Zenit', 'Gris Nube HG', 'Olmo HG']
colors = []
for i, color in enumerate(cName['Name']):
    colors.append(color)

codes =[]
for i, code in enumerate(cName['Code']) :
    codes.append(code)

skus = []
for i, sku in enumerate(pSku['SKU']):
    skus.append(sku)

tPs =[]
for i, tp in enumerate(pSku['T']):
    tPs.append(tp)

ePs =[]
for i, ep in enumerate(pSku['E']):
    ePs.append(ep)

bPs =[]
for i, bp in enumerate(pSku['B']):
    bPs.append(bp)

a1Ps =[]
for i, a1p in enumerate(pSku['A1']):
    a1Ps.append(a1p)

a2Ps =[]
for i, a2p in enumerate(pSku['A2']):
    a2Ps.append(a2p)

a3Ps =[]
for i, a3p in enumerate(pSku['A3']):
    a3Ps.append(a3p)


# 将字典写入到 Excel 文件中,我们使用 openpyxl 库将这个字典写入到一个新的 Excel 文件中，其中第一列包含第一个文件中的值，第二列包含第二个文件中的整个列。
workbook = Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='Handle')
worksheet.cell(row=1, column=2, value='Title')
worksheet.cell(row=1, column=3, value='Option1 Name')
worksheet.cell(row=1, column=4, value='Option1 Value') 
worksheet.cell(row=1, column=5, value='Variant SKU') 
worksheet.cell(row=1, column=6, value='Variant Price') 



skuRow =2 
codeRow =2
for i,sku in enumerate(skus):
    worksheet.cell(row=skuRow, column=2, value=sku)

    for j, color in enumerate(colors):
        worksheet.cell(row=skuRow, column=4, value=color)
        worksheet.cell(row=skuRow,column=1,value=sku)
        worksheet.cell(row=skuRow,column=3,value="Material")
        skuRow +=1
    for x, code in enumerate(codes):
        worksheet.cell(row=codeRow,column=5,value=sku + "-" + code)
        codeRow +=1

# T 的价格是2-10， E:11-18, B: 19-30 , a1：31-34, a2: 35-38, a3:39-41 ; 下一次SKU就是第一次价格出现+40
tRow = 2
for t, tp in enumerate(tPs):
    for i in range(tRow,tRow+9):
        worksheet.cell(row=i, column=6, value=tp)  
    tRow +=40
 
eRow = 11
for t, ep in enumerate(ePs):
    for i in range(eRow,eRow+8):
        worksheet.cell(row=i, column=6, value=ep)
    eRow +=40

bRow = 19
for t, bp in enumerate(bPs):
    for i in range(bRow,bRow+12):
        worksheet.cell(row=i, column=6, value=bp)
    bRow += 40

a1Row = 31
for t, a1p in enumerate(a1Ps):
    for i in range(a1Row,a1Row+4):
        worksheet.cell(row=i, column=6, value=a1p)
    a1Row += 40

a2Row = 35
for t, a2p in enumerate(a2Ps):
    for i in range(a2Row,a2Row+4):
        worksheet.cell(row=i, column=6, value=a2p)
    a2Row += 40

a3Row = 39
for t, a3p in enumerate(a3Ps):
    for i in range(a3Row,a3Row+3):
        worksheet.cell(row=i, column=6, value=a3p)
    a3Row += 40





workbook.save('output.xlsx')
