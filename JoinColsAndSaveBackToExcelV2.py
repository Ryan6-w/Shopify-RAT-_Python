import pandas as pd
from openpyxl import Workbook

color = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Color info_detail.xlsx')
sku = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Cabinet_detailxlsx.xlsx')

# 读取第一个 Excel 文件，提取指定列的数据
cName = pd.read_excel(color, sheet_name='All', usecols=['Name','Code'])

# 读取第二个 Excel 文件，提取指定列的数据
pSku= pd.read_excel(sku, sheet_name='Test', usecols=['SKU','T','E','B','A1','A2','A3'])

colors = ['Brushed Aluminum', 'River Rock', 'Sheer Beauty', 'Fashionista', 'The Chameleon', 'Weekend Getaway', 'Winter Fun', 'Casting at First Light', 'Sugar on Ice', 'Sand Gladstone Oak', 'Grey-Beige Gladstone Oak', 'Brown Tossini Elm', 'Tobacco Gladstone Oak', 'Tobacco Halifax Oak', 'Black Halifax Oak', 'Natural Halifax Oak', 'White Halifax Oak', 'Pearl White HG', 'Winter Frost SM', 'Sun Grey HG', 'Sun Grey SM', 'Stone Grey HG', 'Stone Grey SM', 'Eclipse  HG', 'Eclipse  SM', 'Royal Blue HG', 'Royal Blue SM', 'Majestic HG', 'Majestic SM', 'Ida 01', 'Ida 03', 'Roble Muratti 01', 'Roble Muratti 04', 'Factory 01', 'Factory 02', 'Como Ash 01', 'Como Ash 03', 'Gris Nube Zenit', 'Gris Nube HG', 'Olmo HG']
skus = []
for i, sku in enumerate(pSku['SKU']):
    skus.append(sku)

# 将字典写入到 Excel 文件中,我们使用 openpyxl 库将这个字典写入到一个新的 Excel 文件中，其中第一列包含第一个文件中的值，第二列包含第二个文件中的整个列。
workbook = Workbook()
worksheet = workbook.active

worksheet.cell(row=1, column=1, value='SKU')
worksheet.cell(row=1, column=2, value='Name')

skuRow =2 
colorRow =0
for i in range(len(skus)):
    worksheet.cell(row=skuRow, column=1, value=skus[i])
    #  i 是外部循环的索引，用于计算每个 key 的行索引。j 是内部循环的索引，用于计算 value 列表中的每个元素的行索引。因此，新行的索引为 i+j+2。
    for j in range(len(colors)):
        colorRow = skuRow
        worksheet.cell(row=colorRow, column=2, value=colors[j])
        skuRow +=1

# Save the new workbook
workbook.save('output.xlsx')