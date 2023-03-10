import pandas as pd
from openpyxl import Workbook
import os


def substring_until_number(s):
    result = ""
    for i in s:
        if s[0].isdigit():
            return "DB"
        if i.isdigit():
            break
        result += i
    return result

colorPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Adroit Stocked Color info.xlsx')
productPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/CNG_Cabinet_ Data.xlsx')
newExcelPath = '/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/output.xlsx'

# remove the excel file and csv file
if os.path.exists(newExcelPath):
    os.remove(newExcelPath)


# 读取第一个 Excel 文件，提取指定列的数据
colors = pd.read_excel(colorPath, usecols=['Color name','Panel Code','Price Level'])
# 读取第二个 Excel 文件，提取指定列的数据
products= pd.read_excel(productPath, sheet_name='demo2', usecols=['CABINET','URL','COMODO_BOX','A','B','C','D','E','F'])

# 指定要获取值的列名列表
colorsExtract = ['Color name','Panel Code','Price Level']
# 创建一个空列表，用于存储提取的值
colorsList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in colors.iterrows():
    # 下面代码是简写的展开版本
    # values =[]
    # for column in colorsExtract:
    #     value = row[column]
    #     values.append(value)
    values = [row[columnHeader] for columnHeader in colorsExtract]
    colorsList.append(values)

# 指定要获取值的列名列表
productExtract = ['CABINET','URL','COMODO_BOX','A','B','C','D','E','F']
# 创建一个空列表，用于存储提取的值
productList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in products.iterrows():
    values = [row[columnHeader] for columnHeader in productExtract]
    productList.append(values)


# 将字典写入到 Excel 文件中,我们使用 openpyxl 库将这个字典写入到一个新的 Excel 文件中，其中第一列包含第一个文件中的值，第二列包含第二个文件中的整个列。
workbook = Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='Handle')
worksheet.cell(row=1, column=2, value='Title')
worksheet.cell(row=1, column=3, value='Option1 Name')
worksheet.cell(row=1, column=4, value='Option1 Value') 
worksheet.cell(row=1, column=5, value='Variant SKU') 
worksheet.cell(row=1, column=6, value='Variant Price') 
worksheet.cell(row=1, column=7, value='Status') 
worksheet.cell(row=1, column=8, value='Variant Inventory Policy') 
worksheet.cell(row=1, column=9, value='Variant Fulfillment Service') 
worksheet.cell(row=1, column=10, value='Variant Requires Shipping') 
worksheet.cell(row=1, column=11, value='Variant Taxable') 
worksheet.cell(row=1, column=12, value='Variant Weight Unit') 
worksheet.cell(row=1, column=13, value='Image Src') 
worksheet.cell(row=1, column=14, value='Image Position') 
worksheet.cell(row=1, column=15, value='Tags') 

# 如果没有价格那么价格是String, 有价格会是float 或者int
# for productRow in productList:
    # print(type(productRow[2]))


insertRow = 2 
price =0
count =0
tag = ""
# productList index: 0=sku, 1= url, 2= box price, 3 = A ,4= B, 5=C, 6=D ,7=E ,8 =F 
# colorsList index: 0=name, 1 =code, 2= price level
for productRow in productList:
    if not isinstance(productRow[2],(int,float)):
        print("Product with empty price is: " + str(productRow[0])) # 不要用Remove 因为list 是有序，会自动向上移,会少读一个产品
        productRow =[]
        count +=1
        continue
    
    if not isinstance(productRow[1],(str)):
        print("Product with price but not photo: " + str(productRow[0])) # 不要用Remove 因为list 是有序，会自动向上移,会少读一个产品
        productRow =[]
        count +=1
        continue
    worksheet.cell(row=insertRow, column=2, value=productRow[0])
    worksheet.cell(row=insertRow,column=7,value="active")
    worksheet.cell(row=insertRow,column=13,value=productRow[1]) 
    
    tagSub= substring_until_number(str(productRow[0]))
    if(tagSub =="W"):
        tag = "Wall Cabinet"
    elif(tagSub == "LW"):
        tag ="Lift Up Door_Wall Cabinet"
    elif(tagSub == "DCW"):
        tag ="Dia Corner Wall"
    elif(tagSub == "PCCW"):
        tag = "Pie Cut Wall"
    elif(tagSub == "BCW"):
        tag = "Blind Wall"
    elif(tagSub == "B"):
        tag = "Base Cabinet"
    elif(tagSub == "SB"):
        tag = "Sink Base"
    elif(tagSub == "FSB"):
        tag = "Farm Sink Base"
    elif(tagSub == "DCB" or tagSub=="DCSB"):
        tag = "Diagonal Base"
    elif(tagSub == "PCCB"):
        tag = "Pie Cut Base"
    elif(tagSub == "DB"):
        tag = "Drawer Base"
    elif(tagSub == "BCB" or tagSub=="BCSB"):
        tag = "Base Blind"
    elif(tagSub == "PC"):
        tag = "Pantry"
    elif(tagSub == "MWB"):
        tag = "Microwave"
    elif(tagSub == "OVPC" or tagSub == "OVB"):
        tag = "Oven"
    elif(tagSub == "KND"):
        tag = "Knee Drawer"
    worksheet.cell(row=insertRow,column=15,value=tag) 


    for colorRow in colorsList:
        worksheet.cell(row=insertRow, column=4, value=colorRow[0])
        worksheet.cell(row=insertRow, column=1, value="Cuppowood-"+ str(productRow[0]))
        worksheet.cell(row=insertRow,column=3,value="Material")

        worksheet.cell(row=insertRow,column=5,value=str(productRow[0])+"-"+str(colorRow[1]))
        if(colorRow[2] == 'A'):
            price = round(productRow[2]+productRow[3],2)
        elif (colorRow[2] == 'B'):
            price = round(productRow[2]+productRow[4],2)
        elif (colorRow[2] == 'C'):
            price = round(productRow[2]+productRow[5],2)
        elif (colorRow[2] == 'D'):
            price = round(productRow[2]+productRow[6],2)
        elif (colorRow[2] == 'E'):
            price = round(productRow[2]+productRow[7],2)
        elif (colorRow[2] == 'F'):
            price = round(productRow[2]+productRow[8],2)
        else:
            price =0
        worksheet.cell(row=insertRow,column=6,value= price)
        worksheet.cell(row=insertRow,column=8,value="deny")
        worksheet.cell(row=insertRow,column=9,value="manual")
        worksheet.cell(row=insertRow,column=10,value="TRUE")
        worksheet.cell(row=insertRow,column=11,value="TRUE")
        worksheet.cell(row=insertRow,column=12,value="g")
        insertRow +=1

print("Total removed numbers are: "+ str(count))
workbook.save(newExcelPath)
