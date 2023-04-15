import pandas as pd
from openpyxl import Workbook
import os
import re
import csv
from openpyxl import load_workbook

def seperateSKU(s):
    numString = ""
    for i in s:
        if i.isnumeric():
            numString+=i
            
    return numString

def tagFormat(s):
    return s.replace(' ', '_').replace('(','').replace(')','').replace('\"','').lower()

def info(w,h,d,sku):
    w = str(w)
    h = str(h)
    d = str(d)
    tempTitle = f"{w}\"W {h}\"H {d}\"D ({sku})"
    tempTag = f"{w}W, {h}H, D{d}D"
    pDes = "Width:"+ w +", Height:"+ h +  ", Depth:"+ d  
    return tempTitle,tempTag,pDes

# colorPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Adroit Stocked Color info.xlsx')
# productPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/CNG_Cabinet_ Data.xlsx')
# newExcelPath = '/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/priceAdjustAcc.xlsx'

colorPath = pd.ExcelFile("D:\Spaice\文件\Adroit Stocked Color info.xlsx")
productPath = pd.ExcelFile('D:\Spaice\文件\CNG_Cabinet_ Data.xlsx')
newExcelPath = 'D:\Spaice\output\PriceAdjustAcc.xlsx'

# remove the excel file and csv file
if os.path.exists(newExcelPath):
    os.remove(newExcelPath)

# 读取第一个 Excel 文件，提取指定列的数据
colors = pd.read_excel(colorPath, usecols=['Color name','Panel Code','Price Level'])
# 读取第二个 Excel 文件，提取指定列的数据
products= pd.read_excel(productPath, sheet_name='Acc', usecols=['CABINET','SKU','COMODO_BOX','A','B','C','D','E','F'])

# 指定要获取值的列名列表
colorsExtract = ['Color name','Panel Code','Price Level']
# 创建一个空列表，用于存储提取的值
colorsList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in colors.iterrows():
    values = [row[columnHeader] for columnHeader in colorsExtract]
    colorsList.append(values)

# 指定要获取值的列名列表
productExtract = ['CABINET','SKU','COMODO_BOX','A','B','C','D','E','F']
# 创建一个空列表，用于存储提取的值
productList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in products.iterrows():
    values = [row[columnHeader] for columnHeader in productExtract]
    productList.append(values)


# 将字典写入到 Excel 文件中,我们使用 openpyxl 库将这个字典写入到一个新的 Excel 文件中，其中第一列包含第一个文件中的值，第二列包含第二个文件中的整个列。
workbook = Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='Handle')
worksheet.cell(row=1, column=2, value='Title')
worksheet.cell(row=1, column=3, value='Option1 Name')
worksheet.cell(row=1, column=4, value='Option1 Value') 
worksheet.cell(row=1, column=5, value='Variant Price') 



insertRow = 2 
price = actualPrice = count = depth = height = width = 0
pTitle = pTag = pType = pDes = tempSKU= photoLink = varLink= photoName= colorName= ""


# productList index: 0=sku, 1=kiSKU, 2= box price, 3 = A ,4= B, 5=C, 6=D ,7=E ,8 =F ,
# colorsList index: 0=name, 1 =code, 2= price level
for productRow in productList:

    tempSKU = productRow[1]
    if productRow[0][0:3] != 'DWP': 
        numString =seperateSKU(productRow[0])
    
    # DWP size 需要再问一下
    if productRow[0][0:3] == 'DWP':
        pType = "Panel"

        width = 29.875
        height = 34.5
        depth = 3/4
        tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])

        if productRow[0] =="DWP":
            tempType = "Dishwasher "+pType
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
        elif productRow[0] =="DWP_2DB LOOK":
            tempType = "Dishwasher Panel (2 Drawer Base Look)"
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
        elif productRow[0] =="DWP_3DB LOOK":
            tempType = "Dishwasher Panel (3 Drawer Base Look)"
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
        elif productRow[0] =="DWP_4DB LOOK":
            tempType = "Dishwasher Panel (4 Drawer Base Look)"
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
        elif productRow[0] =="DWP_B1 LOOK":
            tempType = "Dishwasher Panel (B1 Look)"
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
        elif productRow[0] =="DWP_B2 LOOK":
            tempType = "Dishwasher Panel (B2 Look)"
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"

    elif productRow[0][1:4] == "COL":
        pType = "Column"
        
        if productRow[0][0]== 'B':
            tempType = "Base "+pType

            width = numString[0]
            height = 34.5 
            depth = 24     
            
            tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
            
        elif productRow[0][0] == 'W':
            tempType = "Wall "+ pType

            width = numString[0]
            height = numString[1:3]
            depth = numString[3:5]

            tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"

    elif tempSKU[-6:] == "FILLER":
        pType = "Filler"
        
        if productRow[0][:2]== "BF" or productRow[0][:3]== "BLF":
            tempType =  "Base "+pType

            width = numString[0]
            height = numString[1:3]
            if productRow[0][:2]== "BF":
                depth = 3/4
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            else:
                depth = 3
                tempType =  "Base L "+ pType
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            
        elif productRow[0][:2]== "WF" or productRow[0][:3]== "WLF":
            tempType =  "Wall "+pType

            width = numString[0]
            height = numString[1:3]
            if productRow[0][:2]== "WF":
                depth = 3/4
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            else:
                depth = 3
                tempType =  "Wall L "+ pType
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"

        elif productRow[0][:2]== "TF" or productRow[0][:3]== "TLF":
            tempType =  "Tall "+pType

            width = numString[0]
            height = numString[1:3]
            if productRow[0][:2]== "TF":
                depth = 3/4
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            else:
                depth = 3
                tempType =  "Tall L "+pType
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"

    elif tempSKU[-5:] == "PANEL":
        pType = "Panel"
        
        if productRow[0][0]== "B":
            tempType =  "Base "+ pType

            width = numString[:2] 
            height = numString[2:4]
            if height == "35":
                height = "34.5"
            depth = 3/4

            # need to verify the range 
            if width != "35":
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            else:
                tempType = "Back " + pType
                tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"

        # need to verify the range 
        elif productRow[0][0]== "W":
            tempType =  "Wall "+pType

            width = numString[:2]
            height = numString[2:4]
            depth = 3/4

            tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"

        elif productRow[0][:3]== "DWR":
            tempType =  "Dishwasher Return "+pType

            width = numString[0]
            height = 34.5
            depth = 24

            tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
    
        elif productRow[0][:3]== "REP":
            tempType =  "Refrigerator "+pType

            width = numString[:2]
            height = numString[2:4]
            depth = 3/4

            tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
    
    elif tempSKU[-2:] == "TK":
        tempType = pType= "Toe Kick"
        width = 96
        height = 4.5
        depth = 3/4
        tempTitle,tempTag,pDes=info(width,height,depth,productRow[0])
        pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
        pTitle = f"{tempType} {tempTitle}"



    worksheet.cell(row=insertRow, column=2, value=pTitle)
    # worksheet.cell(row=insertRow,column=7,value=photoLink) 

    for colorRow in colorsList:
        if(colorRow[2] == 'A' or colorRow[2] == 'B'):
            colorName = colorRow[0] + " [Classic]"
        elif (colorRow[2] == 'C' or colorRow[2] == 'D'):
            colorName = colorRow[0]+ " [Allure]"
        elif (colorRow[2] == 'E'):
            colorName = colorRow[0] + " [Royal]"
        elif (colorRow[2] == 'F'):
            colorName = colorRow[0] + " [Luxe]"
        worksheet.cell(row=insertRow, column=4, value=colorName)
        worksheet.cell(row=insertRow, column=1, value="Cuppowood-"+ str(productRow[0]))
        worksheet.cell(row=insertRow,column=3,value="Material")

        if(colorRow[2] == 'A'):
            price = round(productRow[3],2)
            # actualPrice = round(price *0.4,2)
        elif (colorRow[2] == 'B'):
            price = round(productRow[4],2)
            # actualPrice = round(price *0.4,2)
        elif (colorRow[2] == 'C'):
            price = round(productRow[5],2)
            # actualPrice = round(price *0.4,2)
        elif (colorRow[2] == 'D'):
            price = round(productRow[6],2)
            # actualPrice = round(price *0.4,2)
        elif (colorRow[2] == 'E'):
            price = round(productRow[7],2)
            # actualPrice = round(price *0.4,2)
        elif (colorRow[2] == 'F'):
            price = round(productRow[8],2)
            # actualPrice = round(price *0.4,2)
        else:
            price =0
        
        worksheet.cell(row=insertRow,column=5,value= price)
        insertRow +=1

print("Total removed numbers are: "+ str(count))
workbook.save(newExcelPath)


newCSVpath = 'D:\Spaice\output\PriceAdjustAcc.csv'

if os.path.exists(newCSVpath):
    os.remove(newCSVpath)

# 读取Excel文件
df = pd.read_excel(newExcelPath)

# 将数据写入CSV文件
df.to_csv(newCSVpath, index=False)