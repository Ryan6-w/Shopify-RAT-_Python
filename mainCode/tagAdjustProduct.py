import pandas as pd
from openpyxl import Workbook
import os
import re
import csv
from openpyxl import load_workbook

def seperateSKU(s):
    numString = ""
    for i in s:
        if i.isnumeric():
            numString+=i
            
    return numString

def tagFormat(s):
    return s.replace(' ', '_').replace('(','').replace(')','').replace('\"','').replace('-',"_").lower()


colorPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Adroit Stocked Color info.xlsx')
productPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/CNG_Cabinet_ Data.xlsx')
newExcelPath = '/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/tagAdjustProduct.xlsx'


# remove the excel file and csv file
if os.path.exists(newExcelPath):
    os.remove(newExcelPath)

# 读取第一个 Excel 文件，提取指定列的数据
colors = pd.read_excel(colorPath, usecols=['Color name','Panel Code','Price Level'])
# 读取第二个 Excel 文件，提取指定列的数据
products= pd.read_excel(productPath, sheet_name='demo1', usecols=['CABINET','SKU','COMODO_BOX','A','B','C','D','E','F'])

# 指定要获取值的列名列表
colorsExtract = ['Color name','Panel Code','Price Level']
# 创建一个空列表，用于存储提取的值
colorsList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in colors.iterrows():
    values = [row[columnHeader] for columnHeader in colorsExtract]
    colorsList.append(values)

# 指定要获取值的列名列表
productExtract = ['CABINET','SKU','COMODO_BOX','A','B','C','D','E','F']
# 创建一个空列表，用于存储提取的值
productList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in products.iterrows():
    values = [row[columnHeader] for columnHeader in productExtract]
    productList.append(values)


# 将字典写入到 Excel 文件中,我们使用 openpyxl 库将这个字典写入到一个新的 Excel 文件中，其中第一列包含第一个文件中的值，第二列包含第二个文件中的整个列。
workbook = Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='Handle')
worksheet.cell(row=1, column=2, value='Title')
worksheet.cell(row=1, column=3, value='Tags')


insertRow = 2 
price = actualPrice=count = depth = height = width = 0
pTitle = pTag = pType = pDes = tempSKU= photoLink = varLink= ""


# productList index: 0=sku, 1=kiSKU, 2= box price, 3 = A ,4= B, 5=C, 6=D ,7=E ,8 =F ,
# colorsList index: 0=name, 1 =code, 2= price level
for productRow in productList:
    if not isinstance(productRow[2],(int,float)):
        # print("Product with empty price is: " + str(productRow[0])) # 不要用Remove 因为list 是有序，会自动向上移,会少读一个产品
        productRow =[]
        count +=1
        continue

    # if not isinstance(productRow[1],(str)):
    if not isinstance(productRow[1],(str)) or productRow[1] == '-':
        # print("Product with price but not photo: " + str(productRow[0])) # 不要用Remove 因为list 是有序，会自动向上移,会少读一个产品
        productRow =[]
        count +=1
        continue
    
    tempSKU = productRow[1]
    if productRow[0][0].isnumeric():
        numString = seperateSKU(productRow[0][1:])
    else:
        numString =seperateSKU(productRow[0])
    
    # 如果要改下面的，还需要改其他2个地方，一个是base cabinet 一个是knee drawer
    width = numString[:2]
    height = numString[2:4]
    depth = numString[4:6]
    tempTitle = f"{width}\"W {height}\"H {depth}\"D ({productRow[0]})"
    tempTag = f"width:{width}, height:{height}, depth:{depth}"
    pDes = "Width:"+ width +", Height:"+ height+  ", Depth:"+ depth  

    if tempSKU[3:5] == "EW":
        pType = "Wall Cabinet"
        
        if tempSKU[3:6]== "EWR":
            if int(depth) ==24:
                tempType = "Refrigerator Wall Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif 30<=int(height)<=42:
                tempType = "High Wall Cabinet"
                pTag = f"{tagFormat(pType)}:{height}_{tagFormat(tempType)}, {tempTag}"
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif int(height)<30:
                tempType = "Standard Hight Wall Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}"
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif 48<=int(height):
                tempType = "Standing Wall Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}"
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"

        elif tempSKU[3:6] == "EWL":
            #K2,HX, HK
            mainType = "Lift Up Door Wall Cabinet"
            if tempSKU[-2:] =="K2":
                tempType = "Standard Lift Up Door Wall Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}"
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:] =="HX":
                tempType = "Lift Up Door Wall Cabinet HK-XS"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}"
                pTitle = f"{tempType} With HK-XS {tempTitle}"
            elif tempSKU[-2:] =="HK":
                tempType = "Lift Up Door Wall Cabinet HK-Top"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}"
                pTitle = f"{tempType} With HK-Top {tempTitle}"

        elif tempSKU[3:6] == "EWC":
            #DR, PR
            if tempSKU[-2:] =="DR":
                tempType = "Diagonal Corner Wall Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}"
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:] =="PR":
                tempType = "Pie Cut Corner Wall Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}"
                pTitle = f"{tempType} {tempTitle}"
                
        elif tempSKU[3:6] == "EWB":
            tempType = "Blind Corner Wall Cabinet"
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}"
            pTitle = f"{tempType} {tempTitle}"
        # print(f"tag: {pTag}")
        # print(f"title: {pTitle}")

    elif(productRow[1][3:5] == "EP"):
        pType = "Pantry"

        if tempSKU[-2:]!= "OV":
            tempType = f"{depth}\" Deep Pantry"
            if tempSKU[-2:] == "PT":
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:] == "R3":
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}_3_ro, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:] == "R4":
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}_4_ro, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:] == "FD":
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}_full_height_door, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} (Full Height Door) {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} (Full Height Door) {tempTitle}_DoubleDoor"
        elif tempSKU[-2:] =="OV":
            tempType = "Oven Pantry"
            pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
            pTitle = f"{tempType} {tempTitle}"
        
        # print(f"tag: {pTag}")
        # print(f"title: {pTitle}")

    elif(productRow[1][3:5] == "EB"):
        pType = "Base Cabinet"
        width = numString[:2]
        height = "34.5"
        depth = "24"
        tempTitle = f"{width}\"W {height}\"H {depth}\"D ({productRow[0]})"
        tempTag = f"width:{width}, height:{height}, depth:{depth}"
        pDes = "Width:"+ width +", Height:"+ height+  ", Depth:"+ depth  
        if tempSKU[3:6]== "EBD":
            tempType = "Drawer Base Cabinet"
            if tempSKU[-2:] == "W1":
                pTag = f"{tagFormat(pType)}:1_{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"1 {tempType} {tempTitle}"
            elif tempSKU[-2:] == "W2":
                pTag = f"{tagFormat(pType)}:2_{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"2 {tempType} {tempTitle}"
            elif tempSKU[-2:] == "T1":
                pTag = f"{tagFormat(pType)}:2_{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"2 {tempType} (1 Top Roll Out Tray) {tempTitle}"
            elif tempSKU[-2:] == "W3":
                pTag = f"{tagFormat(pType)}:3_{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"3 {tempType} {tempTitle}"
            elif tempSKU[-2:] == "W4":
                pTag = f"{tagFormat(pType)}:4_{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"4 {tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBC":
            tempType = "Corner Base Cabinet"
            if tempSKU[-2:] in ("DR","SR"):
                pTag = f"{tagFormat(pType)}:diagonal_{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"Diagonal {tempType} {tempTitle}"
            elif tempSKU[-2:] in ("PR","PW","PM"):
                pTag = f"{tagFormat(pType)}:pie_cut_{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"Pie-Cut {tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBB":
            if tempSKU[-2:]== "FD":
                tempType = "Blind Base Cabinet (Full Height Door)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "BB":
                tempType = "Blind Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} (1 Drawer) {tempTitle}"
            elif tempSKU[-2:]== "SR":
                tempType = "Blind Sink Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "SF":
                tempType = "Blind Sink Base Cabinet (Full Height Door)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBS":
            tempType = "Sink Base Cabinet"
            if tempSKU[-2:]== "BS":
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "S-R1":
                tempType = "Sink Base Cabinet (1 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "S-R2":
                tempType = "Sink Base Cabinet (2 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "TT":
                tempType = "Sink Base Cabinet (Tilt Out)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "FD":
                tempType = "Sink Base Cabinet (Full Height Door)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "FDR1":
                tempType = "Sink Base Cabinet (Full Height Door With Bottom 1 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "FDR2":
                tempType = "Sink Base Cabinet (Full Height Door With Bottom 2 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "FS":
                tempType = "Farm Sink Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBR":
            if tempSKU[-2:]== "BR":
                tempType = "Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:]== "R1":
                tempType = "Base Cabinet (1 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"           
            elif tempSKU[-2:]== "R2":
                tempType = "Base Cabinet (2 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:]== "GP":
                tempType = "Pull-Out Basket Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"            
            elif tempSKU[-2:]== "HM":
                tempType = "Hamper Basket Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "OV":
                tempType = "Oven Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"            
            elif tempSKU[-2:]== "MW":
                tempType = "Microwave Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "KN":
                tempType = "Knee Drawer Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, width:{width}, height:{height}, depth:{depth}" 
                pTitle = f"{tagFormat(pType)} {width}\"W {height}\"H {depth}\" ({productRow[0]})"
        elif tempSKU[3:6]== "EBF":
            if tempSKU[-2:]== "BF":
                tempType = "Base Cabinet (Full Height Door)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:]== "T1":
                tempType = "Base Cabinet (Full Height Door With Top 1 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"  
            elif tempSKU[-2:]== "R1":
                tempType = "Base Cabinet (Full Height Door With Bottom 1 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"        
            elif tempSKU[-2:]== "R2":
                tempType = "Base Cabinet (Full Height Door With 2 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:]== "R3":
                tempType = "Base Cabinet (Full Height Door With 3 Roll Out Tray)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                if int(width) < 24:
                    pTitle = f"{tempType} {tempTitle}_SingleDoor"
                elif int(width) >= 24:
                    pTitle = f"{tempType} {tempTitle}_DoubleDoor"
            elif tempSKU[-2:]== "GP":
                tempType = "Pull-Out Basket Base Cabinet"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "SP":
                tempType = "Pull-Out Basket Base Cabinet (Spice Full Height Door )"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "HM":
                tempType = "Hamper Base Cabinet (Full Height Door)"
                pTag = f"{tagFormat(pType)}:{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"       



    worksheet.cell(row=insertRow, column=1, value="Cuppowood-"+ str(productRow[0]))
    worksheet.cell(row=insertRow, column=2, value=pTitle)
    worksheet.cell(row=insertRow, column=3, value=pTag)
    insertRow +=1

print("Total removed numbers are: "+ str(count))
workbook.save(newExcelPath)


tempCSVPath = '/Users/ryanweng/Documents/Cuppowood/website/产品导入/TagUpdate_template.csv'
newCSVpath = '/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/tagAdjustProduct.csv'

if os.path.exists(newCSVpath):
    os.remove(newCSVpath)


# 打开 csv 文件并读取 header
with open(tempCSVPath, 'r') as f:
    reader = csv.reader(f)
    header = next(reader)

# 打开 Excel 文件
wb = load_workbook(newExcelPath)

# 获取第一个 sheet
ws = wb.active

# 将 Excel 数据读取为 DataFrame
df = pd.DataFrame(ws.values)

# 获取 Excel 数据的 header
excel_header = list(df.iloc[0])

# 将 Excel 数据按照 CSV header 的顺序整理
data = []
for row in df.iloc[1:].values:
    d = {}
    for i, value in enumerate(row):
        d[excel_header[i]] = value
    data.append(d)

# 打开 CSV 文件并写入数据
with open(newCSVpath, 'a', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=header)
    if f.tell() == 0:
        # CSV 文件没有 header，写入 header
        writer.writeheader()
    # 写入数据
    writer.writerows(data)