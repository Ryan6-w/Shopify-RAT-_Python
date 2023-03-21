import pandas as pd
from openpyxl import Workbook
import os
import re

def seperateSKU(s):
    numString = ""
    for i in s:
        if i.isnumeric():
            numString+=i
            
    return numString

def tagFormat(s):
    return s.replace(' ', '_').replace('(','').replace(')','').replace('\"','').lower()


colorPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Adroit Stocked Color info.xlsx')
productPath = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/CNG_Cabinet_ Data.xlsx')
newExcelPath = '/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/output.xlsx'

cURL ="https://s3.us-east-2.amazonaws.com/static.spaice.ca/share/cuppowood/Cabinet/"
cabinetURL ="https://s3.us-east-2.amazonaws.com/static.spaice.ca/share/cuppowood/ConcatPhoto/"

cPhotoPath ="/Users/ryanweng/Documents/Cuppowood/website/产品导入/Shopify/Cabinet/"
cPhotoName = os.listdir(cPhotoPath)

# just to get the photo name is could
cabinetPhotoPath ="/Users/ryanweng/Documents/Cuppowood/website/产品导入/Shopify/ConcatPhoto/"
cabinetPhotoName = os.listdir(cabinetPhotoPath)

# remove the excel file and csv file
if os.path.exists(newExcelPath):
    os.remove(newExcelPath)

# 读取第一个 Excel 文件，提取指定列的数据
colors = pd.read_excel(colorPath, usecols=['Color name','Panel Code','Price Level'])
# 读取第二个 Excel 文件，提取指定列的数据
products= pd.read_excel(productPath, sheet_name='demo2', usecols=['CABINET','SKU','COMODO_BOX','A','B','C','D','E','F'])

# 指定要获取值的列名列表
colorsExtract = ['Color name','Panel Code','Price Level']
# 创建一个空列表，用于存储提取的值
colorsList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in colors.iterrows():
    values = [row[columnHeader] for columnHeader in colorsExtract]
    colorsList.append(values)

# 指定要获取值的列名列表
productExtract = ['CABINET','SKU','COMODO_BOX','A','B','C','D','E','F']
# 创建一个空列表，用于存储提取的值
productList = []
# 遍历每一行，提取指定列的值并添加到列表中；用iterrows 来遍历每一行，index为索引，row 为当前行数
for index, row in products.iterrows():
    values = [row[columnHeader] for columnHeader in productExtract]
    productList.append(values)


# 将字典写入到 Excel 文件中,我们使用 openpyxl 库将这个字典写入到一个新的 Excel 文件中，其中第一列包含第一个文件中的值，第二列包含第二个文件中的整个列。
workbook = Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='Handle')
worksheet.cell(row=1, column=2, value='Title')
worksheet.cell(row=1, column=3, value='Option1 Name')
worksheet.cell(row=1, column=4, value='Option1 Value') 
worksheet.cell(row=1, column=5, value='Variant SKU') 
worksheet.cell(row=1, column=6, value='Variant Price') 
worksheet.cell(row=1, column=7, value='Status') 
worksheet.cell(row=1, column=8, value='Variant Inventory Policy') 
worksheet.cell(row=1, column=9, value='Variant Fulfillment Service') 
worksheet.cell(row=1, column=10, value='Variant Requires Shipping') 
worksheet.cell(row=1, column=11, value='Variant Taxable') 
worksheet.cell(row=1, column=12, value='Variant Weight Unit') 
worksheet.cell(row=1, column=13, value='Image Src') 
worksheet.cell(row=1, column=14, value='Image Position')  # not yet
worksheet.cell(row=1, column=15, value='Tags') 
worksheet.cell(row=1, column=16, value='Product Category') 
worksheet.cell(row=1, column=17, value='Type') 
worksheet.cell(row=1, column=18, value='Body (HTML)') 
worksheet.cell(row=1, column=19, value='Variant Image') 



insertRow = 2 
price = count = depth = height = width = 0
pTitle = pTag = pType = pDes = tempSKU= photoLink = varLink= ""


# productList index: 0=sku, 1=kiSKU, 2= box price, 3 = A ,4= B, 5=C, 6=D ,7=E ,8 =F ,
# colorsList index: 0=name, 1 =code, 2= price level
for productRow in productList:
    if not isinstance(productRow[2],(int,float)):
        # print("Product with empty price is: " + str(productRow[0])) # 不要用Remove 因为list 是有序，会自动向上移,会少读一个产品
        productRow =[]
        count +=1
        continue

    # if not isinstance(productRow[1],(str)):
    if not isinstance(productRow[1],(str)) or productRow[1] == '-':
        # print("Product with price but not photo: " + str(productRow[0])) # 不要用Remove 因为list 是有序，会自动向上移,会少读一个产品
        productRow =[]
        count +=1
        continue
    
    tempSKU = productRow[1]
    if productRow[0][0].isnumeric():
        numString = seperateSKU(productRow[0][1:])
    else:
        numString =seperateSKU(productRow[0])
    
    # 如果要改下面的，还需要改其他2个地方，一个是base cabinet 一个是knee drawer
    width = numString[:2]
    height = numString[2:4]
    depth = numString[4:6]
    tempTitle = f"{width}\"W {height}\"H {depth}\"D ({productRow[0]})"
    tempTag = f"{width}W, {height}H, D{depth}D"
    pDes = "Width:"+ width+ ", Depth:"+ depth +", Height:"+ height 

    if tempSKU[3:5] == "EW":
        pType = "Wall Cabinet"
        
        if tempSKU[3:6]== "EWR":
            if int(depth) ==24:
                tempType = "Refrigerator Wall Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif 30<=int(height)<=42:
                tempType = "High Wall Cabinet"
                pTag = f"{height}_{tagFormat(tempType)}, {tempTag}"
                pTitle = f"{height}\" {tempType} {tempTitle}"
            elif int(height)<30:
                tempType = "Standard Hight Wall Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}"
                pTitle = f"{tempType} {tempTitle}"
            elif 48<=int(height):
                tempType = "Standing Wall Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}"
                pTitle = f"{tempType} {tempTitle}"

        elif tempSKU[3:6] == "EWL":
            #K2,HX, HK
            mainType = "Lift Up Door Wall Cabinet"
            if tempSKU[-2:] =="K2":
                tempType = "Standard Lift Up Door Wall Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, {tagFormat(mainType)}"
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:] =="HX":
                tempType = "Lift Up Door Wall Cabinet HK-XS"
                pTag = f"{tagFormat(tempType)}, {tempTag}, {tagFormat(mainType)}"
                pTitle = f"{tempType} with HK-XS {tempTitle}"
            elif tempSKU[-2:] =="HK":
                tempType = "Lift Up Door Wall Cabinet HK-Top"
                pTag = f"{tagFormat(tempType)}, {tempTag}, {tagFormat(mainType)}"
                pTitle = f"{tempType} with HK-Top {tempTitle}"

        elif tempSKU[3:6] == "EWC":
            #DR, PR
            if tempSKU[-2:] =="DR":
                tempType = "Diagonal Corner Wall Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, diagonal, corner"
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:] =="PR":
                tempType = "Pie Cut Corner Wall Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, pie_cut, corner"
                pTitle = f"{tempType} {tempTitle}"
                
        elif tempSKU[3:6] == "EWB":
            tempType = "Blind Corner Wall Cabinet"
            pTag = f"{tagFormat(tempType)}, {tempTag}, blind, corner"
            pTitle = f"{tempType} {tempTitle}"
        # print(f"tag: {pTag}")
        # print(f"title: {pTitle}")

    elif(productRow[1][3:5] == "EP"):
        pType = "Pantry"

        if tempSKU[-2:]!= "OV":
            tempType = f"{depth}\" Deep Pantry"
            if tempSKU[-2:] == "PT":
                pTag = f"{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:] == "R3":
                pTag = f"{tagFormat(tempType)}, {tempTag}, {tagFormat(tempType)}_3_ro" 
                pTitle = f"{tempType} (3RO) {tempTitle}"
            elif tempSKU[-2:] == "R4":
                pTag = f"{tagFormat(tempType)}, {tempTag}, {tagFormat(tempType)}_4_ro" 
                pTitle = f"{tempType} (4RO) {tempTitle}"
            elif tempSKU[-2:] == "FD":
                pTag = f"{tagFormat(tempType)}, {tempTag}, {tagFormat(tempType)}_fhd, fhd" 
                pTitle = f"{tempType} (FHD) {tempTitle}"
        elif tempSKU[-2:] =="OV":
            tempType = "Oven Pantry"
            pTag = f"{tagFormat(tempType)}, {tempTag}, oven" 
            pTitle = f"{tempType} {tempTitle}"
        
        # print(f"tag: {pTag}")
        # print(f"title: {pTitle}")

    elif(productRow[1][3:5] == "EB"):
        pType = "Base Cabinet"
        width = numString[:2]
        height = "34.5"
        depth = "24"
        tempTitle = f"{width}\"W {height}\"H {depth}\"D ({productRow[0]})"
        tempTag = f"{width}W, {height}H, {depth}D"
        pDes = "Width:"+ width+ ", Depth:"+ depth +", Height:"+ height 
        if tempSKU[3:6]== "EBD":
            tempType = "Drawer Base Cabinet"
            if tempSKU[-2:] == "W1":
                pTag = f"1_{tagFormat(tempType)}, {tagFormat(tempType)}, {tempTag}" 
                pTitle = f"1 {tempType} {tempTitle}"
            elif tempSKU[-2:] == "W2":
                pTag = f"2_{tagFormat(tempType)}, {tagFormat(tempType)}, {tempTag}" 
                pTitle = f"2 {tempType} {tempTitle}"
            elif tempSKU[-2:] == "T1":
                pTag = f"2_{tagFormat(tempType)}, {tagFormat(tempType)}, {tempTag}" 
                pTitle = f"2 {tempType} (Top 1RO) {tempTitle}"
            elif tempSKU[-2:] == "W3":
                pTag = f"3_{tagFormat(tempType)}, {tagFormat(tempType)}, {tempTag}" 
                pTitle = f"3 {tempType} {tempTitle}"
            elif tempSKU[-2:] == "W4":
                pTag = f"4_{tagFormat(tempType)}, {tagFormat(tempType)}, {tempTag}" 
                pTitle = f"4 {tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBC":
            tempType = "Corner Base Cabinet"
            if tempSKU[-2:] in ("DR","SR"):
                pTag = f"diagonal_{tagFormat(tempType)}, {tempTag}, diagonal, corner" 
                pTitle = f"Diagonal {tempType} {tempTitle}"
            elif tempSKU[-2:] in ("PR","PW","PM"):
                pTag = f"pie_cut_{tagFormat(tempType)}, {tempTag}, pie_cut, corner" 
                pTitle = f"Pie-Cut {tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBB":
            if tempSKU[-2:]== "FD":
                tempType = "Blind Base Cabinet (FHD)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "BB":
                tempType = "Blind Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, blind" 
                pTitle = f"{tempType} (1 Drawer) {tempTitle}"
            elif tempSKU[-2:]== "SR":
                tempType = "Blind Sink Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, blind, sink" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "SF":
                tempType = "Blind Sink Base Cabinet (FHD)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, blind, sink" 
                pTitle = f"{tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBS":
            tempType = "Sink Base Cabinet"
            if tempSKU[-2:]== "BS":
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "S-R1":
                tempType = "Sink Base Cabinet (1RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "S-R2":
                tempType = "Sink Base Cabinet (2RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "TT":
                tempType = "Sink Base Cabinet (Tilt Out)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "FD":
                tempType = "Sink Base Cabinet (FHD)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink, fhd" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "FDR1":
                tempType = "Sink Base Cabinet (FHD BOT 1RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink, fhd" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-4:]== "FDR2":
                tempType = "Sink Base Cabinet (FHD BOT 2RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink, fhd" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "FS":
                tempType = "Farm Sink Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, sink" 
                pTitle = f"{tempType} {tempTitle}"
        elif tempSKU[3:6]== "EBR":
            if tempSKU[-2:]== "BR":
                tempType = "Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "R1":
                tempType = "Base Cabinet (1RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"            
            elif tempSKU[-2:]== "R2":
                tempType = "Base Cabinet (2RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "GP":
                tempType = "Pull-Out Basket Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, pull_out_basket" 
                pTitle = f"{tempType} {tempTitle}"            
            elif tempSKU[-2:]== "HM":
                tempType = "Hamper Basket Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, hamper_basket" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "OV":
                tempType = "Oven Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, oven" 
                pTitle = f"{tempType} {tempTitle}"            
            elif tempSKU[-2:]== "MW":
                tempType = "Microwave Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, microwave" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "KN":
                tempType = "Knee Drawer Cabinet"
                pTag = f"{tagFormat(tempType)}, {width}W, {height}H, {depth}D" 
                pTitle = f"{tempType} {width}\"W {height}\"H {depth}\" ({productRow[0]})"
        elif tempSKU[3:6]== "EBF":
            if tempSKU[-2:]== "BF":
                tempType = "Base Cabinet (FHD)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "T1":
                tempType = "Base Cabinet (FHD Top 1RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd" 
                pTitle = f"{tempType} {tempTitle}"   
            elif tempSKU[-2:]== "R1":
                tempType = "Base Cabinet (FHD BOT 1RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd" 
                pTitle = f"{tempType} {tempTitle}"         
            elif tempSKU[-2:]== "R2":
                tempType = "Base Cabinet (FHD 2RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "R3":
                tempType = "Base Cabinet (FHD 3RO)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "GP":
                tempType = "Pull-Out Basket Base Cabinet"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd, pull_out_basket" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "SP":
                tempType = "Pull-Out Basket Base Cabinet (FHD Spice)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd, pull_out_basket" 
                pTitle = f"{tempType} {tempTitle}"
            elif tempSKU[-2:]== "HM":
                tempType = "Hamper Base Cabinet (FHD)"
                pTag = f"{tagFormat(tempType)}, {tempTag}, fhd, hamper" 
                pTitle = f"{tempType} {tempTitle}"       

        # print(f"tag: {pTag}")
        # print(f"title: {pTitle}")

    width = int(width)
    photoSKU = productRow[1][3:].replace('-', '')
    for cName in cPhotoName:
        if photoSKU == "EBFGP" and width == 15:
            pattern1 = re.compile(rf".+-{photoSKU}_SINGLE.jpg")
            if re.match(pattern1, cName):
                photoLink = cURL + cName
        elif photoSKU == "EBFGP" and width == 15:
            pattern1 = re.compile(rf".+-{photoSKU}_DOUBLE.jpg")
            if re.match(pattern1, cName):
                photoLink = cURL + cName
        elif width < 24 :        
            pattern1 = re.compile(rf".+-{photoSKU}(?:_SINGLEDOOR)?.jpg")
            if re.match(pattern1, cName):
                photoLink = cURL + cName
        elif width >= 24 :        
            pattern1 = re.compile(rf".+-{photoSKU}(?:_DOUBLEDOOR)?.jpg")
            if re.match(pattern1, cName):
                photoLink = cURL + cName


    worksheet.cell(row=insertRow, column=2, value=pTitle)
    worksheet.cell(row=insertRow,column=7,value="active")
    worksheet.cell(row=insertRow,column=13,value=photoLink) 
    worksheet.cell(row=insertRow,column=15,value=pTag) 
    worksheet.cell(row=insertRow,column=16,value="Furniture > Cabinets & Storage > Kitchen Cabinets") 
    worksheet.cell(row=insertRow,column=17,value=pType) 
    worksheet.cell(row=insertRow,column=18,value=pDes) 


    for colorRow in colorsList:
        worksheet.cell(row=insertRow, column=4, value=colorRow[0])
        worksheet.cell(row=insertRow, column=1, value="Cuppowood-"+ str(productRow[0]))
        worksheet.cell(row=insertRow,column=3,value="Material")

        worksheet.cell(row=insertRow,column=5,value=str(productRow[0])+"-"+str(colorRow[1]))
        if(colorRow[2] == 'A'):
            price = round(productRow[2]+productRow[3],2)
        elif (colorRow[2] == 'B'):
            price = round(productRow[2]+productRow[4],2)
        elif (colorRow[2] == 'C'):
            price = round(productRow[2]+productRow[5],2)
        elif (colorRow[2] == 'D'):
            price = round(productRow[2]+productRow[6],2)
        elif (colorRow[2] == 'E'):
            price = round(productRow[2]+productRow[7],2)
        elif (colorRow[2] == 'F'):
            price = round(productRow[2]+productRow[8],2)
        else:
            price =0

        
        # width大于或者等于24 为doubledoor, 小于24为Singledoor
        # EBF-GP B15 => 单， B18=> 双
        tempColor = colorRow[0].replace(' ','').replace('-','')
        # .+任意字符串
        # (?P<name>pattern) =》以下语法来创建命名捕获组
        # 使用了非捕获组 (?:_SINGLEDOOR)? 和 (?:_DOUBLEDOOR)?，表示它们是可选的，即可能存在也可能不存在
        for cName in cabinetPhotoName:
            if photoSKU == "EBFGP" and width == 15:
                pattern = re.compile(rf"B_FHD_GPO-EBFGP_SINGLE--{tempColor}")
                if re.match(pattern, cName):
                    varLink = cabinetURL + cName

            elif photoSKU == "EBFGP" and width == 18:
                pattern = re.compile(rf"B_FHD_GPO-EBFGP_DOUBLE--{tempColor}")
                if re.match(pattern, cName):
                    varLink = cabinetURL + cName

            elif width < 24 :        
                pattern = re.compile(rf".+-{photoSKU}(?:_SINGLEDOOR)?--{tempColor}")
                if re.match(pattern, cName):
                    varLink = cabinetURL + cName
            elif width >= 24 :        
                pattern = re.compile(rf".+-{photoSKU}(?:_DOUBLEDOOR)?--{tempColor}")
                if re.match(pattern, cName):
                    varLink = cabinetURL + cName
        

        worksheet.cell(row=insertRow,column=6,value= price)
        worksheet.cell(row=insertRow,column=8,value="deny")
        worksheet.cell(row=insertRow,column=9,value="manual")
        worksheet.cell(row=insertRow,column=10,value="TRUE")
        worksheet.cell(row=insertRow,column=11,value="TRUE")
        worksheet.cell(row=insertRow,column=12,value="g")
        worksheet.cell(row=insertRow,column=19,value=varLink)
        insertRow +=1

print("Total removed numbers are: "+ str(count))
workbook.save(newExcelPath)
