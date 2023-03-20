import csv
import pandas as pd
from openpyxl import Workbook

tempURL ="https://s3.us-east-2.amazonaws.com/static.spaice.ca/share/cuppowood/Color/"
color = pd.ExcelFile('/Users/ryanweng/Documents/Cuppowood/website/产品导入/Adroit Stocked Color info.xlsx')


# 读取第一个 Excel 文件，提取指定列的数据
cName = pd.read_excel(color, usecols=['Color name'])

colors = []
for i, color in enumerate(cName['Color name']):
    colors.append(color)

workbook = Workbook()
worksheet = workbook.active
worksheet.cell(row=1, column=1, value='Value')
worksheet.cell(row=1, column=2, value='Select type')
worksheet.cell(row=1, column=3, value='Custom')


for i,color in enumerate(colors):
    worksheet.cell(row=i+2, column=1, value=color)
    worksheet.cell(row=i+2,column=2,value="Image url")

    url = tempURL+color.replace(" ","+")+".jpg"
    worksheet.cell(row=i+2,column=3,value=url)

workbook.save('/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/color.xlsx')

# # 读取Excel文件
# excel_file = '/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/color.xlsx'
# df = pd.read_excel(excel_file)

# # 保存为CSV文件
# csv_file = '/Users/ryanweng/Documents/Cuppowood/Python/Testfiles/color.csv'
# df.to_csv(csv_file, index=False)